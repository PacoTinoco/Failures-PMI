"""
Router: CO — Changeover Analysis (Análisis de Cambio de Marca)
Módulo para análisis de changeovers en Filtros Cápsula.
Incluye normalización de máquinas/operadores, split multi-operador,
tracking de completitud y exportación a Excel.
"""

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from app.services.supabase_client import get_supabase_admin
import pandas as pd
import unicodedata
import re
from io import BytesIO
from datetime import date

router = APIRouter(prefix="/co", tags=["CO"])


# ══════════════════════════════════════════════════════
# Normalization helpers
# ══════════════════════════════════════════════════════

VALID_KDFS = {7, 8, 9, 10, 11, 17}

def normalize_machine(raw: str) -> str:
    if not raw or str(raw).strip().lower() in ('nan', '', 'na'):
        return None
    s = str(raw).strip().upper()
    s = re.sub(r'[^A-Z0-9]', '', s)  # remove dashes, spaces
    m = re.match(r'KDF(\d+)', s)
    if m:
        num = int(m.group(1))
        if num in VALID_KDFS:
            return f"KDF {num}"
    m2 = re.match(r'MULFI', s, re.IGNORECASE)
    if m2:
        return "KDF 10"
    return str(raw).strip()


def strip_accents(s: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )


OPERATOR_ALIASES = {
    'milron': 'Milton',
    'milton': 'Milton',
    'mílton': 'Milton',
    'monse': 'Monse',
    'nico': 'Nico',
    'eduardo v': 'Eduardo V',
}


def normalize_operator(raw: str) -> str:
    if not raw or str(raw).strip().lower() in ('nan', '', 'na'):
        return None
    s = str(raw).strip()
    s = re.sub(r'\s+', ' ', s)  # collapse multiple spaces
    key = strip_accents(s).lower().strip()
    return OPERATOR_ALIASES.get(key, s)


def split_operators(raw: str) -> list:
    if not raw or str(raw).strip().lower() in ('nan', '', 'na'):
        return []
    s = str(raw).strip()
    if ' y ' in s.lower():
        parts = re.split(r'\s+y\s+', s, flags=re.IGNORECASE)
        return [normalize_operator(p) for p in parts if p.strip()]
    return [normalize_operator(s)]


# ══════════════════════════════════════════════════════
# Models
# ══════════════════════════════════════════════════════

class COCreate(BaseModel):
    maquina: str
    fecha: str
    semana: Optional[str] = None
    operador: str
    marca_termina: Optional[str] = None
    marca_nueva: Optional[str] = None
    formatos_completados: Optional[str] = None
    analisis_cf: Optional[float] = None
    razon_desviacion: Optional[str] = None
    recomendacion: Optional[str] = None
    tiempo_objetivo: Optional[float] = None
    tiempo_real: Optional[float] = None
    runtime_next_2h: Optional[float] = None
    stops_next_2h: Optional[int] = None
    mtbf: Optional[float] = None
    variacion_co: Optional[float] = None
    desperdicio_hora2: Optional[float] = None

class COUpdate(BaseModel):
    maquina: Optional[str] = None
    fecha: Optional[str] = None
    semana: Optional[str] = None
    operador: Optional[str] = None
    marca_termina: Optional[str] = None
    marca_nueva: Optional[str] = None
    formatos_completados: Optional[str] = None
    analisis_cf: Optional[float] = None
    razon_desviacion: Optional[str] = None
    recomendacion: Optional[str] = None
    tiempo_objetivo: Optional[float] = None
    tiempo_real: Optional[float] = None
    runtime_next_2h: Optional[float] = None
    stops_next_2h: Optional[int] = None
    mtbf: Optional[float] = None
    variacion_co: Optional[float] = None
    desperdicio_hora2: Optional[float] = None


# ══════════════════════════════════════════════════════
# CRUD endpoints
# ══════════════════════════════════════════════════════

@router.get("/records")
async def list_co(
    semana: Optional[str] = Query(None),
    maquina: Optional[str] = Query(None),
    operador: Optional[str] = Query(None),
):
    sb = get_supabase_admin()
    query = sb.table("co_records").select("*").order("fecha", desc=True)
    if semana:
        query = query.eq("semana", semana)
    if maquina:
        query = query.eq("maquina", maquina)
    if operador:
        query = query.eq("operador", operador)
    result = query.execute()
    return {"data": result.data}


@router.get("/records/{co_id}")
async def get_co(co_id: str):
    sb = get_supabase_admin()
    result = sb.table("co_records").select("*").eq("id", co_id).execute()
    if not result.data:
        raise HTTPException(404, "CO record not found")
    return {"data": result.data[0]}


@router.post("/records")
async def create_co(body: COCreate):
    sb = get_supabase_admin()
    record = body.dict()
    record["maquina"] = normalize_machine(record["maquina"])
    record["operador"] = normalize_operator(record["operador"])
    # Compute analisis_cf from completeness
    has_razon = bool(record.get("razon_desviacion") and record["razon_desviacion"].strip())
    has_recom = bool(record.get("recomendacion") and record["recomendacion"].strip())
    filled = sum([has_razon, has_recom])
    record["analisis_cf"] = round(filled / 2, 2) if record.get("analisis_cf") is None else record["analisis_cf"]
    result = sb.table("co_records").insert(record).execute()
    return {"data": result.data[0] if result.data else None}


@router.patch("/records/{co_id}")
async def update_co(co_id: str, body: COUpdate):
    sb = get_supabase_admin()
    updates = {k: v for k, v in body.dict().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No fields to update")
    if "maquina" in updates:
        updates["maquina"] = normalize_machine(updates["maquina"])
    if "operador" in updates:
        updates["operador"] = normalize_operator(updates["operador"])
    result = sb.table("co_records").update(updates).eq("id", co_id).execute()
    return {"data": result.data[0] if result.data else None}


@router.delete("/records/{co_id}")
async def delete_co(co_id: str):
    sb = get_supabase_admin()
    sb.table("co_records").delete().eq("id", co_id).execute()
    return {"ok": True}


# ══════════════════════════════════════════════════════
# Statistics & Analysis
# ══════════════════════════════════════════════════════

@router.get("/stats")
async def get_co_stats():
    sb = get_supabase_admin()
    records = sb.table("co_records").select("*").order("fecha").execute().data
    if not records:
        return {
            "total": 0, "by_machine": {}, "by_operator": {},
            "by_week": {}, "avg_variacion": 0, "avg_mtbf": 0,
            "completeness": {}, "brand_pairs": [],
        }

    by_machine = {}
    by_operator = {}
    by_week = {}
    variaciones = []
    mtbfs = []
    completeness = {}  # operator → {total, filled_razon, filled_recom, details[]}
    brand_pairs = {}
    time_data = []
    all_marcas_termina = set()
    all_marcas_nueva = set()

    def _valid_str(val):
        if not val:
            return None
        s = str(val).strip()
        if s.lower() in ('', 'nan', 'na', 'none'):
            return None
        return s

    for r in records:
        maq = r.get("maquina") or "N/A"
        by_machine[maq] = by_machine.get(maq, 0) + 1

        op = r.get("operador") or "N/A"
        by_operator[op] = by_operator.get(op, 0) + 1

        sem = r.get("semana") or "N/A"
        by_week[sem] = by_week.get(sem, 0) + 1

        if r.get("variacion_co") is not None:
            variaciones.append(r["variacion_co"])

        if r.get("mtbf") is not None:
            mtbfs.append(r["mtbf"])

        # Track unique brand names
        mt = _valid_str(r.get("marca_termina"))
        mn = _valid_str(r.get("marca_nueva"))
        if mt:
            all_marcas_termina.add(mt)
        if mn:
            all_marcas_nueva.add(mn)

        # Completeness tracking with detail records
        if op not in completeness:
            completeness[op] = {
                "total": 0, "filled_razon": 0, "filled_recom": 0,
                "filled_desperdicio": 0, "desperdicio_values": [],
                "details": [],
            }
        completeness[op]["total"] += 1

        razon_val = _valid_str(r.get("razon_desviacion"))
        recom_val = _valid_str(r.get("recomendacion"))
        desp_val = r.get("desperdicio_hora2")

        if razon_val:
            completeness[op]["filled_razon"] += 1
        if recom_val:
            completeness[op]["filled_recom"] += 1
        if desp_val is not None:
            completeness[op]["filled_desperdicio"] += 1
            completeness[op]["desperdicio_values"].append(desp_val)

        # Store detail record for dropdown
        completeness[op]["details"].append({
            "fecha": r.get("fecha"),
            "maquina": maq,
            "marca_termina": mt,
            "marca_nueva": mn,
            "razon_desviacion": razon_val,
            "recomendacion": recom_val,
            "desperdicio_hora2": desp_val,
            "variacion_co": r.get("variacion_co"),
        })

        # Brand pair tracking
        pair_key = f"{mt or '?'} → {mn or '?'}"
        brand_pairs[pair_key] = brand_pairs.get(pair_key, 0) + 1

        # Time deviation data
        if r.get("tiempo_real") is not None and r.get("tiempo_objetivo") is not None:
            time_data.append({
                "fecha": r["fecha"],
                "semana": sem,
                "maquina": maq,
                "operador": op,
                "objetivo": r["tiempo_objetivo"],
                "real": r["tiempo_real"],
                "variacion": r.get("variacion_co"),
            })

    # Compute completeness percentages + avg desperdicio
    for op in completeness:
        t = completeness[op]["total"]
        completeness[op]["pct_razon"] = round(completeness[op]["filled_razon"] / t * 100) if t else 0
        completeness[op]["pct_recom"] = round(completeness[op]["filled_recom"] / t * 100) if t else 0
        completeness[op]["pct_desperdicio"] = round(completeness[op]["filled_desperdicio"] / t * 100) if t else 0
        dvals = completeness[op]["desperdicio_values"]
        completeness[op]["avg_desperdicio"] = round(sum(dvals) / len(dvals), 2) if dvals else None
        del completeness[op]["desperdicio_values"]  # no need to send raw list

    top_brands = sorted(brand_pairs.items(), key=lambda x: -x[1])[:15]

    return {
        "total": len(records),
        "by_machine": by_machine,
        "by_operator": by_operator,
        "by_week": by_week,
        "avg_variacion": round(sum(variaciones) / len(variaciones), 2) if variaciones else 0,
        "avg_mtbf": round(sum(mtbfs) / len(mtbfs), 2) if mtbfs else 0,
        "completeness": completeness,
        "brand_pairs": [{"pair": k, "count": v} for k, v in top_brands],
        "time_data": time_data,
        "total_operators": len(set(r.get("operador") for r in records if r.get("operador"))),
        "total_machines": len(set(r.get("maquina") for r in records if r.get("maquina"))),
        "weeks_range": sorted(set(r.get("semana") for r in records if r.get("semana"))),
        "all_marcas_termina": sorted(all_marcas_termina),
        "all_marcas_nueva": sorted(all_marcas_nueva),
    }


# ══════════════════════════════════════════════════════
# Seed from Excel
# ══════════════════════════════════════════════════════

COLUMN_MAP = {
    'máquina': 'maquina', 'maquina': 'maquina',
    'fecha': 'fecha',
    'semana': 'semana',
    'eo': 'operador',
    'marca que termina': 'marca_termina',
    'marca nueva': 'marca_nueva',
    'formatos completados': 'formatos_completados',
    'analisis cf': 'analisis_cf',
    'razón de desviación': 'razon_desviacion', 'razon de desviacion': 'razon_desviacion',
    'recomendación de mejoras': 'recomendacion', 'recomendacion de mejoras': 'recomendacion',
    'tiempo objetivo': 'tiempo_objetivo',
    'tiempo real': 'tiempo_real',
    'run time next 2 hours': 'runtime_next_2h',
    '#stops next 2 hours': 'stops_next_2h',
    'mtbf': 'mtbf',
    'variación de co': 'variacion_co', 'variacion de co': 'variacion_co',
    'desperdicio en la hora 2': 'desperdicio_hora2',
}


@router.post("/seed")
async def seed_from_excel(file: UploadFile = File(...), clear_existing: bool = Query(False)):
    contents = await file.read()
    df = pd.read_excel(BytesIO(contents))

    # Map columns
    col_mapping = {}
    for orig_col in df.columns:
        key = str(orig_col).strip().lower()
        if key in COLUMN_MAP:
            col_mapping[orig_col] = COLUMN_MAP[key]
    df = df.rename(columns=col_mapping)

    # Only keep columns we know
    known_cols = set(COLUMN_MAP.values())
    df = df[[c for c in df.columns if c in known_cols]]

    sb = get_supabase_admin()

    if clear_existing:
        existing = sb.table("co_records").select("id").execute().data
        for r in existing:
            sb.table("co_records").delete().eq("id", r["id"]).execute()

    inserted = 0
    skipped = 0
    split_count = 0
    errors = []

    for idx, row in df.iterrows():
        try:
            raw_op = str(row.get('operador', '')).strip()
            operators = split_operators(raw_op)
            if not operators:
                skipped += 1
                continue

            # Parse fecha
            fecha = None
            raw_fecha = row.get('fecha')
            if pd.notna(raw_fecha):
                try:
                    fecha = pd.to_datetime(raw_fecha).strftime('%Y-%m-%d')
                except:
                    fecha = str(raw_fecha)

            # Parse semana
            semana = None
            raw_sem = row.get('semana')
            if pd.notna(raw_sem):
                try:
                    semana = str(int(raw_sem))
                except:
                    semana = str(raw_sem).strip()

            # Numeric fields
            def safe_float(val):
                if pd.isna(val):
                    return None
                try:
                    return float(val)
                except:
                    return None

            def safe_int(val):
                if pd.isna(val):
                    return None
                try:
                    return int(float(val))
                except:
                    return None

            def safe_str(val):
                if pd.isna(val) or str(val).strip().lower() in ('nan', 'na', 'none', ''):
                    return None
                return str(val).strip()

            base_record = {
                "maquina": normalize_machine(str(row.get('maquina', ''))),
                "fecha": fecha,
                "semana": semana,
                "marca_termina": safe_str(row.get('marca_termina')),
                "marca_nueva": safe_str(row.get('marca_nueva')),
                "formatos_completados": safe_str(row.get('formatos_completados')),
                "analisis_cf": safe_float(row.get('analisis_cf')),
                "razon_desviacion": safe_str(row.get('razon_desviacion')),
                "recomendacion": safe_str(row.get('recomendacion')),
                "tiempo_objetivo": safe_float(row.get('tiempo_objetivo')),
                "tiempo_real": safe_float(row.get('tiempo_real')),
                "runtime_next_2h": safe_float(row.get('runtime_next_2h')),
                "stops_next_2h": safe_int(row.get('stops_next_2h')),
                "mtbf": safe_float(row.get('mtbf')),
                "variacion_co": safe_float(row.get('variacion_co')),
                "desperdicio_hora2": safe_float(row.get('desperdicio_hora2')),
            }

            if len(operators) > 1:
                split_count += 1

            for op in operators:
                record = {**base_record, "operador": op}
                sb.table("co_records").insert(record).execute()
                inserted += 1

        except Exception as e:
            errors.append(f"Fila {idx + 2}: {str(e)}")
            skipped += 1

    return {
        "message": f"Seed completado: {inserted} registros insertados",
        "inserted": inserted,
        "skipped": skipped,
        "split_multi_operator": split_count,
        "errors": errors[:10],
    }


# ══════════════════════════════════════════════════════
# Export to Excel
# ══════════════════════════════════════════════════════

@router.get("/export")
async def export_co_excel(
    semana: Optional[str] = Query(None),
    maquina: Optional[str] = Query(None),
    operador: Optional[str] = Query(None),
):
    sb = get_supabase_admin()
    query = sb.table("co_records").select("*").order("fecha", desc=True)
    if semana:
        query = query.eq("semana", semana)
    if maquina:
        query = query.eq("maquina", maquina)
    if operador:
        query = query.eq("operador", operador)

    records = query.execute().data
    if not records:
        raise HTTPException(404, "No hay registros CO para exportar")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis CO"

    headers = [
        "Máquina", "Fecha", "Semana", "Operador",
        "Marca Termina", "Marca Nueva", "Formatos",
        "Análisis CF", "Razón Desviación", "Recomendación",
        "T. Objetivo", "T. Real", "Variación CO",
        "Runtime 2h", "Stops 2h", "MTBF", "Desperdicio H2"
    ]
    col_widths = [10, 12, 8, 16, 16, 16, 10, 11, 30, 30, 11, 11, 12, 11, 10, 10, 14]

    header_fill = PatternFill("solid", fgColor="1F3864")
    thin = Side(border_style="thin", color="7F7F7F")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value="Análisis CO — Changeover Analysis")
    c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")

    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Headers
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all

    # Data
    for row_idx, r in enumerate(records, start=4):
        vals = [
            r.get("maquina"), r.get("fecha"), r.get("semana"), r.get("operador"),
            r.get("marca_termina"), r.get("marca_nueva"), r.get("formatos_completados"),
            r.get("analisis_cf"), r.get("razon_desviacion"), r.get("recomendacion"),
            r.get("tiempo_objetivo"), r.get("tiempo_real"), r.get("variacion_co"),
            r.get("runtime_next_2h"), r.get("stops_next_2h"), r.get("mtbf"),
            r.get("desperdicio_hora2"),
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.font = Font(name="Calibri", size=10)
            cell.border = border_all
            cell.alignment = Alignment(wrap_text=True) if col in (9, 10) else Alignment(horizontal="center")

    ws.freeze_panes = "A4"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=CO_Analysis_Export.xlsx"}
    )
