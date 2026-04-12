"""
Router: IPS — Issue Problem Solving
Dashboard interactivo para registro, seguimiento y contramedidas de IPS.
"""

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from app.services.supabase_client import get_supabase_admin
import pandas as pd
from io import BytesIO
from datetime import date

router = APIRouter(prefix="/ips", tags=["IPS"])


# ══════════════════════════════════════════════════════
# Models
# ══════════════════════════════════════════════════════

class IPSCreate(BaseModel):
    cedula_id: str
    kdf: int
    titulo: str
    fecha: Optional[str] = None
    ubicacion: Optional[str] = None
    participants: Optional[List[str]] = []
    section_6w2h: bool = False
    section_bbc: bool = False
    section_5w: bool = False
    section_res: bool = False
    status: str = "Open"
    priority: Optional[str] = None
    notes: Optional[str] = None

class IPSUpdate(BaseModel):
    kdf: Optional[int] = None
    titulo: Optional[str] = None
    fecha: Optional[str] = None
    ubicacion: Optional[str] = None
    participants: Optional[List[str]] = None
    section_6w2h: Optional[bool] = None
    section_bbc: Optional[bool] = None
    section_5w: Optional[bool] = None
    section_res: Optional[bool] = None
    status: Optional[str] = None
    priority: Optional[str] = None
    notes: Optional[str] = None

class CMCreate(BaseModel):
    ips_id: str
    descripcion: str
    owner: Optional[str] = None
    status: str = "Pending"
    priority: Optional[str] = None
    due_date: Optional[str] = None
    notes: Optional[str] = None

class CMUpdate(BaseModel):
    descripcion: Optional[str] = None
    owner: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[str] = None
    due_date: Optional[str] = None
    notes: Optional[str] = None


# ══════════════════════════════════════════════════════
# IPS CRUD
# ══════════════════════════════════════════════════════

@router.get("/records")
async def list_ips(cedula_id: str = Query(...)):
    sb = get_supabase_admin()
    result = sb.table("ips_records") \
        .select("*") \
        .eq("cedula_id", cedula_id) \
        .order("fecha", desc=True) \
        .execute()
    return {"data": result.data}


@router.get("/records/{ips_id}")
async def get_ips(ips_id: str):
    sb = get_supabase_admin()
    result = sb.table("ips_records").select("*").eq("id", ips_id).execute()
    if not result.data:
        raise HTTPException(404, "IPS not found")
    return {"data": result.data[0]}


@router.post("/records")
async def create_ips(body: IPSCreate):
    sb = get_supabase_admin()
    record = body.dict()
    result = sb.table("ips_records").insert(record).execute()
    return {"data": result.data[0] if result.data else None}


@router.patch("/records/{ips_id}")
async def update_ips(ips_id: str, body: IPSUpdate):
    sb = get_supabase_admin()
    updates = {k: v for k, v in body.dict().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No fields to update")
    result = sb.table("ips_records").update(updates).eq("id", ips_id).execute()
    return {"data": result.data[0] if result.data else None}


@router.delete("/records/{ips_id}")
async def delete_ips(ips_id: str):
    sb = get_supabase_admin()
    # Countermeasures cascade-delete via FK
    sb.table("ips_records").delete().eq("id", ips_id).execute()
    return {"ok": True}


# ══════════════════════════════════════════════════════
# Countermeasures CRUD
# ══════════════════════════════════════════════════════

@router.get("/countermeasures")
async def list_countermeasures(ips_id: str = Query(...)):
    sb = get_supabase_admin()
    result = sb.table("ips_countermeasures") \
        .select("*") \
        .eq("ips_id", ips_id) \
        .order("created_at") \
        .execute()
    return {"data": result.data}


@router.get("/countermeasures/all")
async def list_all_countermeasures(cedula_id: str = Query(...)):
    """Get all countermeasures for all IPS of a cedula (for dashboard stats)."""
    sb = get_supabase_admin()
    # Get IPS IDs first
    ips_records = sb.table("ips_records") \
        .select("id") \
        .eq("cedula_id", cedula_id) \
        .execute().data
    if not ips_records:
        return {"data": []}
    ips_ids = [r["id"] for r in ips_records]

    BATCH = 50
    all_cms = []
    for i in range(0, len(ips_ids), BATCH):
        batch = ips_ids[i:i + BATCH]
        result = sb.table("ips_countermeasures") \
            .select("*") \
            .in_("ips_id", batch) \
            .order("created_at") \
            .limit(5000) \
            .execute()
        all_cms.extend(result.data)
    return {"data": all_cms}


@router.post("/countermeasures")
async def create_countermeasure(body: CMCreate):
    sb = get_supabase_admin()
    result = sb.table("ips_countermeasures").insert(body.dict()).execute()
    return {"data": result.data[0] if result.data else None}


@router.patch("/countermeasures/{cm_id}")
async def update_countermeasure(cm_id: str, body: CMUpdate):
    sb = get_supabase_admin()
    updates = {k: v for k, v in body.dict().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No fields to update")
    result = sb.table("ips_countermeasures").update(updates).eq("id", cm_id).execute()
    return {"data": result.data[0] if result.data else None}


@router.delete("/countermeasures/{cm_id}")
async def delete_countermeasure(cm_id: str):
    sb = get_supabase_admin()
    sb.table("ips_countermeasures").delete().eq("id", cm_id).execute()
    return {"ok": True}


# ══════════════════════════════════════════════════════
# Dashboard stats
# ══════════════════════════════════════════════════════

@router.get("/stats")
async def get_stats(cedula_id: str = Query(...)):
    sb = get_supabase_admin()
    records = sb.table("ips_records") \
        .select("id, kdf, status, ubicacion, fecha, priority") \
        .eq("cedula_id", cedula_id) \
        .execute().data

    if not records:
        return {"total": 0, "by_status": {}, "by_kdf": {}, "by_ubicacion": {}}

    ips_ids = [r["id"] for r in records]
    # Countermeasures stats
    BATCH = 50
    all_cms = []
    for i in range(0, len(ips_ids), BATCH):
        batch = ips_ids[i:i + BATCH]
        cms = sb.table("ips_countermeasures") \
            .select("ips_id, status") \
            .in_("ips_id", batch) \
            .limit(5000) \
            .execute().data
        all_cms.extend(cms)

    by_status = {}
    by_kdf = {}
    by_ubicacion = {}
    for r in records:
        s = r["status"] or "Open"
        by_status[s] = by_status.get(s, 0) + 1
        k = str(r["kdf"])
        by_kdf[k] = by_kdf.get(k, 0) + 1
        u = r["ubicacion"] or "N/A"
        by_ubicacion[u] = by_ubicacion.get(u, 0) + 1

    cm_status = {}
    for cm in all_cms:
        s = cm["status"] or "Pending"
        cm_status[s] = cm_status.get(s, 0) + 1

    return {
        "total": len(records),
        "by_status": by_status,
        "by_kdf": by_kdf,
        "by_ubicacion": by_ubicacion,
        "total_countermeasures": len(all_cms),
        "cm_by_status": cm_status,
    }


# ══════════════════════════════════════════════════════
# Dedup — Remove duplicate countermeasures
# ══════════════════════════════════════════════════════

@router.post("/dedup")
async def dedup_countermeasures(cedula_id: str = Query(...)):
    """Remove duplicate countermeasures (same ips_id + descripcion)."""
    sb = get_supabase_admin()
    ips_records = sb.table("ips_records") \
        .select("id") \
        .eq("cedula_id", cedula_id) \
        .execute().data
    if not ips_records:
        return {"removed": 0}

    ips_ids = [r["id"] for r in ips_records]
    BATCH = 50
    all_cms = []
    for i in range(0, len(ips_ids), BATCH):
        batch = ips_ids[i:i + BATCH]
        cms = sb.table("ips_countermeasures") \
            .select("id, ips_id, descripcion, created_at") \
            .in_("ips_id", batch) \
            .order("created_at") \
            .limit(5000) \
            .execute().data
        all_cms.extend(cms)

    # Find duplicates: keep first occurrence, delete rest
    seen = {}  # (ips_id, descripcion_lower) → first id
    to_delete = []
    for cm in all_cms:
        key = (cm["ips_id"], cm["descripcion"].strip().lower())
        if key in seen:
            to_delete.append(cm["id"])
        else:
            seen[key] = cm["id"]

    for cm_id in to_delete:
        sb.table("ips_countermeasures").delete().eq("id", cm_id).execute()

    return {"removed": len(to_delete), "message": f"Se eliminaron {len(to_delete)} contramedidas duplicadas"}


# ══════════════════════════════════════════════════════
# Export Excel — IPS records + countermeasures
# ══════════════════════════════════════════════════════

@router.get("/export")
async def export_ips_excel(
    cedula_id: str = Query(...),
    kdf: Optional[int] = Query(None),
):
    """Export IPS records and countermeasures to Excel."""
    sb = get_supabase_admin()

    query = sb.table("ips_records") \
        .select("*") \
        .eq("cedula_id", cedula_id) \
        .order("kdf") \
        .order("fecha", desc=True)
    if kdf is not None:
        query = query.eq("kdf", kdf)

    records = query.execute().data
    if not records:
        raise HTTPException(404, "No hay registros IPS para exportar")

    ips_ids = [r["id"] for r in records]
    BATCH = 50
    all_cms = []
    for i in range(0, len(ips_ids), BATCH):
        batch = ips_ids[i:i + BATCH]
        cms = sb.table("ips_countermeasures") \
            .select("*") \
            .in_("ips_id", batch) \
            .order("created_at") \
            .limit(5000) \
            .execute().data
        all_cms.extend(cms)

    # Group countermeasures by IPS id
    cms_by_ips = {}
    for cm in all_cms:
        cms_by_ips.setdefault(cm["ips_id"], []).append(cm)

    # Build print-ready workbook using openpyxl directly
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    wb = Workbook()
    ws = wb.active
    ws.title = "IPS"

    # Column widths (7 columns: KDF, Fecha, Título / Contramedida, Ubicación / Owner, Participantes / Status CM, Status IPS / Prioridad, Fecha límite)
    col_widths = [8, 13, 42, 22, 28, 14, 13]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin = Side(border_style="thin", color="7F7F7F")
    medium = Side(border_style="medium", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_heavy = Border(left=medium, right=medium, top=medium, bottom=medium)

    header_fill = PatternFill("solid", fgColor="1F3864")
    ips_fill = PatternFill("solid", fgColor="D9E1F2")
    cm_fill = PatternFill("solid", fgColor="F2F2F2")
    title_fill = PatternFill("solid", fgColor="2E75B6")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Title row ──
    title_text = f"IPS — Issue Problem Solving" + (f"  |  KDF {kdf}" if kdf else "")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_widths))
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # Blank row
    row_idx = 3

    # ── Header row ──
    headers = ["KDF", "Fecha", "Título / Contramedida", "Ubicación / Owner",
               "Participantes / Status CM", "Status IPS / Prioridad", "Fecha límite"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border_heavy
    ws.row_dimensions[row_idx].height = 30
    row_idx += 1

    # ── Data rows: one IPS row + nested CM rows ──
    for r in records:
        # IPS row
        ips_values = [
            f"KDF {r['kdf']}",
            str(r.get("fecha") or ""),
            r.get("titulo") or "",
            r.get("ubicacion") or "",
            ", ".join(r.get("participants", []) or []),
            r.get("status") or "",
            "",
        ]
        for col, v in enumerate(ips_values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.font = Font(name="Calibri", size=11, bold=True, color="1F3864")
            cell.fill = ips_fill
            cell.alignment = left if col in (3, 4, 5) else center
            cell.border = border_all
        ws.row_dimensions[row_idx].height = 26
        row_idx += 1

        # Nested CM rows
        cms_list = cms_by_ips.get(r["id"], [])
        if not cms_list:
            cell = ws.cell(row=row_idx, column=3, value="— Sin contramedidas registradas —")
            cell.font = Font(name="Calibri", size=10, italic=True, color="7F7F7F")
            cell.fill = cm_fill
            cell.alignment = left
            for col in range(1, len(col_widths) + 1):
                c2 = ws.cell(row=row_idx, column=col)
                c2.fill = cm_fill
                c2.border = border_all
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
        else:
            for idx, cm in enumerate(cms_list, start=1):
                cm_values = [
                    "",
                    "",
                    f"   ↳ {idx}. {cm.get('descripcion') or ''}",
                    cm.get("owner") or "",
                    cm.get("status") or "",
                    cm.get("priority") or "",
                    str(cm.get("due_date") or ""),
                ]
                for col, v in enumerate(cm_values, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=v)
                    cell.font = Font(name="Calibri", size=10, color="333333")
                    cell.fill = cm_fill
                    cell.alignment = left if col in (3, 4) else center
                    cell.border = border_all
                ws.row_dimensions[row_idx].height = 22
                row_idx += 1

        # Empty separator row between IPS blocks
        row_idx += 1

    # Freeze header
    ws.freeze_panes = f"A{4}"

    # Print setup: landscape, fit to width, margins
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3)
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"{3}:{3}"  # repeat header row on each printed page

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"IPS_Export_KDF{kdf}.xlsx" if kdf else "IPS_Export.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ══════════════════════════════════════════════════════
# Upload Excel — Parse and import IPS + Countermeasures
# ══════════════════════════════════════════════════════

@router.post("/upload")
async def upload_ips_excel(
    cedula_id: str = Query(...),
    file: UploadFile = File(...)
):
    """Parse IPS Excel file and import records + countermeasures."""
    contents = await file.read()
    xls = pd.ExcelFile(BytesIO(contents))
    sb = get_supabase_admin()

    imported_ips = 0
    imported_cm = 0
    skipped = 0
    duplicates_skipped = 0

    # ── Load existing IPS for this cedula (to avoid duplicates) ──
    existing_ips = sb.table("ips_records") \
        .select("id, kdf, titulo") \
        .eq("cedula_id", cedula_id) \
        .execute().data
    existing_ips_map = {}  # (kdf, titulo) → id
    for r in existing_ips:
        existing_ips_map[(r["kdf"], r["titulo"])] = r["id"]

    # ── Parse IPS sheet ──
    ips_sheet = None
    for name in xls.sheet_names:
        if 'ips' in name.lower():
            ips_sheet = name
            break
    if not ips_sheet:
        raise HTTPException(400, "No se encontró una hoja con 'IPS' en el nombre")

    df_ips = pd.read_excel(xls, sheet_name=ips_sheet, header=None)
    headers = df_ips.iloc[0].tolist()
    df_data = df_ips.iloc[2:].copy()
    df_data.columns = headers

    df_data = df_data[df_data['KDF'].notna() & (df_data['KDF'] != 'KDF')].copy()

    ips_map = {}  # (kdf, titulo) → ips_id for linking countermeasures
    for _, row in df_data.iterrows():
        try:
            kdf = int(row['KDF'])
        except (ValueError, TypeError):
            skipped += 1
            continue

        titulo = str(row.get('Titulo', '')).strip()
        if not titulo or titulo == 'nan':
            skipped += 1
            continue

        # Skip if IPS already exists (dedup by kdf+titulo)
        if (kdf, titulo) in existing_ips_map:
            ips_map[(kdf, titulo)] = existing_ips_map[(kdf, titulo)]
            duplicates_skipped += 1
            continue

        participants = []
        for col in ['P1','P2','P3','P4','P5','P6','P7','P8']:
            val = row.get(col)
            if pd.notna(val) and str(val).strip() and str(val).strip() not in ['P1','P2','P3','P4','P5','P6','P7','P8']:
                participants.append(str(val).strip())

        fecha = None
        if pd.notna(row.get('Fecha')):
            try:
                fecha = pd.to_datetime(row['Fecha']).strftime('%Y-%m-%d')
            except:
                pass

        ubicacion = str(row.get('Ubicacion', '')).strip() if pd.notna(row.get('Ubicacion')) else None
        if ubicacion in ['', 'nan', 'Ubi', 'Ubicacion']:
            ubicacion = None

        status = str(row.get('O/C', 'Open')).strip()
        if status in ['nan', '', 'O/C']:
            status = 'Open'

        record = {
            "cedula_id": cedula_id,
            "kdf": kdf,
            "titulo": titulo,
            "fecha": fecha,
            "ubicacion": ubicacion,
            "participants": participants,
            "section_6w2h": row.get('6W2H') == 'X',
            "section_bbc": row.get('BBC') == 'X',
            "section_5w": row.get('5W') == 'X',
            "section_res": row.get('Res') == 'X',
            "status": status,
        }

        result = sb.table("ips_records").insert(record).execute()
        if result.data:
            ips_id = result.data[0]["id"]
            ips_map[(kdf, titulo)] = ips_id
            imported_ips += 1

    # ── Parse Countermeasures sheet ──
    cm_sheet = None
    for name in xls.sheet_names:
        if 'contram' in name.lower():
            cm_sheet = name
            break

    if cm_sheet:
        df_cm = pd.read_excel(xls, sheet_name=cm_sheet, header=None)
        cm_headers = df_cm.iloc[0].tolist()
        df_cm_data = df_cm.iloc[2:].copy()
        df_cm_data.columns = cm_headers

        df_cm_data = df_cm_data[df_cm_data['Contramedidas'].notna()].copy()
        df_cm_data = df_cm_data[df_cm_data['Contramedidas'].astype(str).str.strip() != '']

        df_cm_data['KDF'] = df_cm_data['KDF'].ffill()
        df_cm_data['Titulo'] = df_cm_data['Titulo'].ffill()

        # ── Load existing CMs to deduplicate ──
        all_ips_ids = list(set(ips_map.values()) | set(existing_ips_map.values()))
        existing_cms = set()  # (ips_id, descripcion) tuples
        BATCH = 50
        for i in range(0, len(all_ips_ids), BATCH):
            batch = all_ips_ids[i:i + BATCH]
            cms = sb.table("ips_countermeasures") \
                .select("ips_id, descripcion") \
                .in_("ips_id", batch) \
                .limit(5000) \
                .execute().data
            for cm in cms:
                existing_cms.add((cm["ips_id"], cm["descripcion"].strip().lower()))

        for _, row in df_cm_data.iterrows():
            try:
                kdf = int(row['KDF'])
            except (ValueError, TypeError):
                continue

            titulo = str(row.get('Titulo', '')).strip()
            if titulo in ['', 'nan', 'Titulo']:
                continue

            descripcion = str(row['Contramedidas']).strip()
            if not descripcion or descripcion == 'nan':
                continue

            owner = str(row.get('Owner', '')).strip() if pd.notna(row.get('Owner')) else None
            if owner in ['', 'nan']:
                owner = None

            cm_status = str(row.get('Status', 'Pending')).strip()
            if cm_status in ['', 'nan', 'Status']:
                cm_status = 'Pending'

            priority = str(row.get('Priority', '')).strip() if pd.notna(row.get('Priority')) else None
            if priority in ['', 'nan']:
                priority = None

            # Find matching IPS record
            ips_id = ips_map.get((kdf, titulo))
            if not ips_id:
                # Try existing map
                ips_id = existing_ips_map.get((kdf, titulo))
            if not ips_id:
                # Try partial match
                for (k, t), iid in {**ips_map, **existing_ips_map}.items():
                    if k == kdf and (t.startswith(titulo[:20]) or titulo.startswith(t[:20])):
                        ips_id = iid
                        break

            if not ips_id:
                skipped += 1
                continue

            # Skip if this CM already exists (dedup by ips_id + descripcion)
            if (ips_id, descripcion.strip().lower()) in existing_cms:
                duplicates_skipped += 1
                continue

            cm_record = {
                "ips_id": ips_id,
                "descripcion": descripcion,
                "owner": owner,
                "status": cm_status,
                "priority": priority,
            }
            sb.table("ips_countermeasures").insert(cm_record).execute()
            existing_cms.add((ips_id, descripcion.strip().lower()))
            imported_cm += 1

    return {
        "message": f"Importados {imported_ips} IPS y {imported_cm} contramedidas ({duplicates_skipped} duplicados omitidos)",
        "imported_ips": imported_ips,
        "imported_cm": imported_cm,
        "skipped": skipped,
        "duplicates_skipped": duplicates_skipped,
    }
